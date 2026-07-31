"""Ingest the vendored precedent xlsm through the SAME Phase-1 substrate rules use (D-PREC).

The D-PREC audit (02-PRECEDENT-AUDIT.md) is COMPLETE and its dedupe/forward-fill/row-identity
policy is DECIDED -- this module is a MECHANICAL implementation of that policy, not a re-audit.
Precedent chunks are supporting evidence only, never a finding source (D-RB2(5)); no agent tool
is registered here (D-RB3(b) -- tool exposure is a Phase-3-evidence-gated decision, deferred like
D-RB3's own precedent-search deferral).
"""
from __future__ import annotations

import hashlib
import sqlite3
from pathlib import Path

import openpyxl

from ingest.anchors import mint_span
from ingest.normalize import normalize
from ingest.serialize import SERIALIZER_VERSION, serialize_document
from rulebook.store import RuleChunk

_SHEET_NAME = "CMC Def. RoadMap"
_HEADER_ROW = 2   # row 1 = title, row 2 = header, row 3+ = data (02-PRECEDENT-AUDIT.md)
_DB_PATH = "data/defpredict.db"   # SAME db file store.py/edges.py each independently point at

# On-sheet header text -> internal field name (02-PRECEDENT-AUDIT.md's 9-column schema). Read by
# NAME off row 2, not by position -- robust to column reordering in the vendored file. The
# on-sheet text for cols 6/7 is the FULL "... of Deficiency" form (verified against the actual
# committed rulebook/precedents/*.xlsm), not the audit doc's shorthand "Cohort Year"/"Category".
_HEADER_MAP = {
    "ANDA #": "anda_number", "Product Name": "product_name", "Dosage Form": "dosage_form",
    "CMC Section": "cmc_section", "Deficiency Type": "deficiency_type",
    "Cohort Year of Deficiency": "cohort_year", "Category of Deficiency": "category",
    "Deficiency": "deficiency_text", "Deficiency Response": "deficiency_response",
}


def _precedent_row_id(anda_number: str, row_ordinal: int, deficiency_text: str) -> str:
    """D-PREC row-identity formula -- stable across rebuilds, independent of sheet renumbering."""
    key = f"{anda_number}|{row_ordinal}|{deficiency_text}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def _read_rows(xlsm_path: str | Path) -> list[dict]:
    """openpyxl read_only + data_only -- values only, no formulas/formatting."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    ws = wb[_SHEET_NAME]
    header = next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))
    fields = [_HEADER_MAP.get(str(h).strip()) if h else None for h in header]
    rows = []
    for row_ordinal, raw in enumerate(ws.iter_rows(min_row=_HEADER_ROW + 1, values_only=True), start=1):
        row = {f: v for f, v in zip(fields, raw) if f is not None}
        if not row.get("deficiency_text"):
            continue   # a genuinely empty row -- not a blank-ANDA row, skip
        row["row_ordinal"] = row_ordinal
        rows.append(row)
    wb.close()
    return rows


def _forward_fill(rows: list[dict]) -> list[dict]:
    """Forward-fill blank anda_number/product_name/dosage_form from the nearest non-blank row
    above (spreadsheet merge semantics) and stamp anda_inferred -- D-PREC policy item 3. Exactly
    83 rows are expected to need this at ingestion time."""
    last = {"anda_number": None, "product_name": None, "dosage_form": None}
    for row in rows:
        row["anda_inferred"] = not row.get("anda_number")
        for field in last:
            if row.get(field):
                last[field] = row[field]
            else:
                row[field] = last[field]
    return rows


def _group_by_exact_text(rows: list[dict]) -> dict[str, list[dict]]:
    """Exact-text dedupe at chunk level -- D-PREC policy item 2. No near-dup/semantic dedupe."""
    groups: dict[str, list[dict]] = {}
    for row in rows:
        groups.setdefault(row["deficiency_text"], []).append(row)
    return groups


def _ensure_provenance_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        "CREATE TABLE IF NOT EXISTS precedent_provenance "
        "(precedent_row_id TEXT PRIMARY KEY, doc_id TEXT, anda_number TEXT, product_name TEXT, "
        "dosage_form TEXT, cmc_section TEXT, deficiency_type TEXT, cohort_year TEXT, "
        "category TEXT, anda_inferred INTEGER)"
    )


def _write_provenance(doc_id: str, group_rows: list[dict], db_path: str = _DB_PATH) -> None:
    """The provenance-row LIST for one deduped chunk. RuleChunk (Plan 02-02) has no list-of-rows
    field, so provenance lives in its OWN table keyed by doc_id -- mirrors store.py/edges.py's
    own-table-per-concern convention rather than changing store.py's schema."""
    conn = sqlite3.connect(db_path); conn.row_factory = sqlite3.Row
    _ensure_provenance_table(conn)
    for row in group_rows:
        row_id = _precedent_row_id(row.get("anda_number") or "", row["row_ordinal"], row["deficiency_text"])
        conn.execute(
            "INSERT INTO precedent_provenance (precedent_row_id, doc_id, anda_number, product_name, "
            "dosage_form, cmc_section, deficiency_type, cohort_year, category, anda_inferred) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(precedent_row_id) DO UPDATE SET doc_id=excluded.doc_id, "
            "anda_number=excluded.anda_number, product_name=excluded.product_name, "
            "dosage_form=excluded.dosage_form, cmc_section=excluded.cmc_section, "
            "deficiency_type=excluded.deficiency_type, cohort_year=excluded.cohort_year, "
            "category=excluded.category, anda_inferred=excluded.anda_inferred",
            (row_id, doc_id, row.get("anda_number"), row.get("product_name"), row.get("dosage_form"),
             row.get("cmc_section"), row.get("deficiency_type"), str(row.get("cohort_year") or ""),
             row.get("category"), int(row["anda_inferred"])),
        )
    conn.commit(); conn.close()


def get_provenance(doc_id: str, db_path: str = _DB_PATH) -> list[dict]:
    """Read back the provenance-row LIST for one deduped precedent chunk."""
    conn = sqlite3.connect(db_path); conn.row_factory = sqlite3.Row
    _ensure_provenance_table(conn)
    rows = conn.execute("SELECT * FROM precedent_provenance WHERE doc_id = ?", (doc_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _persist_chunk(doc_dict: dict, doc_id: str, citation: str, store) -> RuleChunk:
    """The SAME substrate call rulebook.build._ingest_and_persist makes (serialize_document ->
    normalize -> mint_span -> RuleChunk -> store.write_chunk) -- genuinely reused primitives, not
    a parallel canonicalization path -- parameterized by `store` so ingest_precedents(xlsm_path,
    store) matches its documented signature and tests can target an isolated backend."""
    raw, _cell_ranges = serialize_document(doc_dict)
    nt = normalize(raw, serializer_version=SERIALIZER_VERSION)
    span = mint_span(nt.canonical, 0, len(nt.canonical), doc_id, nt.normalizer_version)
    chunk = RuleChunk(doc_id=doc_id, citation=citation, source="precedent", version="vendored as-is",
                      license="internal", url="", span=span, normalizer_version=nt.normalizer_version,
                      serializer_version=nt.serializer_version)
    store.write_chunk(chunk, nt)
    return chunk


def ingest_precedents(xlsm_path: str | Path, store=None) -> list[RuleChunk]:
    """Parse the vendored precedent xlsm (Task 3's `rulebook/precedents/...xlsm`), apply the
    D-PREC forward-fill + exact-dedupe policy, and persist one RuleChunk per DISTINCT
    deficiency_text -- each carrying its full provenance-row list in this module's own table.
    `store` defaults to the real `rulebook.store` module (has `.write_chunk`); tests may pass an
    isolated double. Never touches Databricks -- local build only (precedent-search Databricks
    serving, if ever built, is Plan 02-08 territory, and only if that plan chooses to extend).
    """
    if store is None:
        import rulebook.store as store
    rows = _forward_fill(_read_rows(xlsm_path))
    groups = _group_by_exact_text(rows)
    chunks: list[RuleChunk] = []
    for deficiency_text, group_rows in groups.items():
        doc_id = "precedent-" + hashlib.sha256(deficiency_text.encode("utf-8")).hexdigest()[:16]
        # anda_number comes back from openpyxl as int (the sheet stores it as a numeric column,
        # not text) -- str() it before joining, since '21 CFR'-style citation strings and this
        # SQLite-agnostic join both expect str, not the raw Excel cell type.
        andas = sorted({str(r.get("anda_number")) if r.get("anda_number") else "UNKNOWN" for r in group_rows})
        citation = f"Precedent deficiency ({len(group_rows)} occurrence(s) across ANDA {', '.join(andas)})"
        doc_dict = {
            "filename": "ANDA-TDDS-Deficiency-Roadmap.xlsm", "page_count": 1, "toc": [],
            "pages": [{
                "page_number": 1, "page_label": "", "width": 612.0, "height": 792.0,
                "rotation": 0, "source": "precedent-xlsm", "is_scanned": False,
                "blocks": [{"text": deficiency_text, "page": 1, "reading_order": 0, "lines": []}],
                "tables": [], "figures": [],
            }],
        }
        chunk = _persist_chunk(doc_dict, doc_id, citation, store)
        _write_provenance(doc_id, group_rows)
        chunks.append(chunk)
    return chunks
